import openpyxl
import os
import shutil
from pprint import pprint

def getExcelFile(file):
    # 读取清单文档
    wb = openpyxl.load_workbook(file)
    # 获取所有sheets
    sheets = wb.sheetnames
    fileDic = {}
    allFiles = []
    for i in range(len(sheets)):
        filelist = []
        sheet = wb[sheets[i]]
        rows = sheet.max_row
        col = 1

        for j in range(rows):
            row = j + 1
            filelist.append(sheet.cell(row, col).value)
            allFiles.append(sheet.cell(row, col).value)

        fileDic[sheets[i]] = filelist
    return fileDic,allFiles

def create_dir_structure(path):
    all = ['eGovaMobile（标准智信apk需改打包地址不包含市政环保环卫多网）',
           'eUrbanMIS',
           '其他',
           '更新说明（必读）.txt']
    all_other = ['多网', '环保', '市政园林采集员环卫排水']
    for i in range(len(all)):
        if i == len(all) - 1:
            f = open(path + '\\' + all[i], 'w')
            f.close()
        elif i == len(all) - 2:
            os.mkdir(path + '\\' + all[i])
            for j in all_other:
                os.mkdir(path + '\\' + all[i] + '\\' + j)
        else:
            os.mkdir(path + '\\' + all[i])

def beforeProcess(path):
    # 删除eUrbanMIS下view/mobile文件夹
    if os.path.exists(path + '\\eUrbanMIS\\view\\mobile'):
        shutil.rmtree(path + '\\eUrbanMIS\\view\\mobile')

    # 删除eUrbanMIS/WEB-INF/lib/下的egova-statis-ex.jar
    file_dir = path + '\\eUrbanMIS\\WEB-INF\\lib'
    if os.path.exists(file_dir + '\\egova-statis-ex-1.0.1-mysql.jar'):
        os.remove(file_dir + '\\egova-statis-ex-1.0.1-mysql.jar')

def getLibPackage(path, fileDic, allFiles ,hotfixLibPath, headers = ['lib','其他','多网','环保','市政园林采集员环卫排水']):
    libPath = path+'\\eUrbanMIS\\WEB-INF\\lib'
    libFiles = os.listdir(libPath)

    for i in libFiles:
        if i in allFiles:
            for j in headers:
                if i in fileDic[j]:
                    namePath = libPath + '\\' + i
                    tempPath = path+'\\'+j
                    if j == 'lib':
                        pass
                    elif j == '其他':
                        shutil.copy(namePath, tempPath)
                        os.remove(namePath)
                    else:
                        tempPath = path+'\\其他\\'+j
                        shutil.copy(namePath, tempPath)
                        os.remove(namePath)
        else:
            print(i+'为新增的包或者包名已修改，请仔细检查！！！')
            with open(path+'\\'+'log.txt', 'a') as f:
                f.write(i+'为新增的包或者包名已修改，请仔细检查！\n')

    for i in allFiles:
        if i not in libFiles:
            if 'egova' not in i:
                try:
                    shutil.copy(hotfixLibPath+'\\'+i, libPath)
                except:
                    print(i + f'在{hotfixLibPath}中不存在，请仔细检查！！！')
                    with open(path + '\\' + 'log.txt', 'a') as f:
                        f.write(i + f'在{hotfixLibPath}中不存在，请仔细检查！\n')
            else:
                print(i+'被修改名称或者没做修改，请仔细检查！！！')
                with open(path + '\\' + 'log.txt', 'a') as f:
                    f.write(i + '被修改名称或者没做修改，请仔细检查！\n')

def get_files(path,files_list=[],dir_list=[]):
    all = os.listdir(path)
    items = path.split('\\')

    for i in range(len(path.split('\\'))):
        num = len(items)-i-1
        if items[num] != '':
            for i in all:
                str = path + '\\' + i
                if os.path.isdir(str):
                    dir_list.append(path)
                else:
                    file_dict=str
                    files_list.append(file_dict)
            break
        else:
            continue

    for i in all:
        str = path+'\\'+i
        if os.path.isdir(str):
            get_files(str, files_list, dir_list)

    return files_list

def get_jsinfo(path):
    '''
    :param path: 根路径
    :return:
    '''
    file_list = get_files(path)
    with open(path+'\\修改的JS文件记录.txt','w') as i:
        for f in file_list:
            if f.endswith('.js'):
                i.write(f+'\n')
        i.write('\n'+'重点JS文件：'+'\n')
        for f in file_list:
            if f.endswith('i18n.js'):
                i.write(f + '\n')

def work_main(file):
    path = input('请输入根目录：\n')
    if os.listdir(path):
        print('根目录不为空。\n清理根目录。')
        for i in os.listdir(path):
            item_path = path + '\\' + i
            if os.path.isfile(item_path):
                os.remove(item_path)
            else:
                shutil.rmtree(item_path)
        print('根目录清理完成。')

        create_dir_structure(path)
        print('文件结构完成。')
    else:
        create_dir_structure(path)
        print('文件结构完成。')

    difFile_flag = input('是否已放入差异文件至eUrbanMIS下(y/n);\n')
    hotfixLibPath = input('请输入hotfix的lib路径：\n')
    beforeProcess(path)

    fileDic, allFiles = getExcelFile(file)

    getLibPackage(path, fileDic,allFiles,hotfixLibPath)

    get_jsinfo(path)

    print("完成！！！")

if __name__ == '__main__':
    file = r'D:\0501版本管理\lib_files.xlsx'
    work_main(file)
    # get_files(r'C:\Users\asd\Desktop\0701')