import openpyxl

def getFileContext(file, sheet_name=''):
    wb = openpyxl.load_workbook(file)
    allSheets = wb.sheetnames
    # print(wb.sheetnames)

    workSheet = wb[sheet_name]

    # 删除不需要的列
    heades = ['端子','光缆名称','二维码','光路编码','业务名称','A端纤芯号']
    index_num = 0
    indexs = []
    for column in workSheet.columns:
        index_num = index_num+1
        if column[0].value in heades:
            indexs.append(index_num)
    # print(indexs)

    context=[]
    for row in workSheet.rows:
        row_context = []
        for i in indexs:
            num = i -1
            row_context.append(row[num].value)
        context.append(row_context)
    wb.close()

    return context


def getNewExcelFile(context):

    # 获取所有数据
    for i in range(len(context)):
        if i == 0:
            context[i].append('备注')
        else:
            context[i].append('')
    print(context)
    rowsNum = len(context)
    columnsNum = 7

    # 获取3列数据，为空的去掉
    sheet1 =  [['二维码','光路编码','业务名称']]
    nums1 = []
    for i in range(columnsNum):
        if context[0][i] in sheet1[0]:
            nums1.append(i)
    # print(nums1)
    for i in range(1,rowsNum):
        temp1 = []
        for j in nums1:
            temp1.append(context[i][j])
        if temp1 == ['', '', '']:
            pass
        else:
            sheet1.append(temp1)
    for i in sheet1:
        if i == '':
            sheet1.remove(i)
    print(sheet1)

    # 宽带业务,第三列标红
    kuandaiSheet = []
    key1 = "宽带"
    kuandaiSheet.append(sheet1[0])
    for data in sheet1:
        try:
            if key1 in data[2]:
                kuandaiSheet.append(data)
        except:
            continue
    print(kuandaiSheet)

    #基站业务，第三列标绿
    jizhanSheet = []
    key2 = "BBU"
    jizhanSheet.append(sheet1[0])
    for data in sheet1:
        try:
            if key2 in data[2]:
                jizhanSheet.append(data)
        except:
            continue
    print(jizhanSheet)

    #PON网业务
    PonSheet = []
    PonSheet.append(sheet1[0])
    for data in sheet1:
        if data not in kuandaiSheet and data not in jizhanSheet:
            PonSheet.append(data)
    print(PonSheet)

    # 基站业务，第三列标绿

    # 宽带业务,第三列标红

    # 所有sheet顶端对齐，自动换行，自动调整行高

    # 调整表格每列的宽度，重新自动换行。

    # 添加排序：主-光路编码，次-A端纤芯号，数值排序

    # 最后保存context、kuandaiSheet、jizhanSheet、PonSheet
    wb = openpyxl.Workbook()
    wb.create_sheet('Context')
    wb.create_sheet('宽带业务')
    wb.create_sheet('基站业务')
    wb.create_sheet('PON网业务')

    ws1 = wb['Context']
    ws2 = wb['宽带业务']
    ws3 = wb['基站业务']
    ws4 = wb['PON网业务']

    for i in range(len(context)):
        for j in range(len(context[0])):
            ws1.cell(row=i+1, column=j+1, value=context[i][j] )

    for i in range(len(kuandaiSheet)):
        for j in range(len(kuandaiSheet[0])):
            ws2.cell(row=i+1, column=j+1, value=kuandaiSheet[i][j] )

    for i in range(len(jizhanSheet)):
        for j in range(len(jizhanSheet[0])):
            ws3.cell(row=i+1, column=j+1, value=jizhanSheet[i][j] )

    for i in range(len(PonSheet)):
        for j in range(len(PonSheet[0])):
            ws4.cell(row=i+1, column=j+1, value=PonSheet[i][j] )

    wb.save('test.xlsx')

    wb2 = openpyxl.load_workbook('test.xlsx')
    sheets =  ['Context', '宽带业务', '基站业务', 'PON网业务']
    for i in wb2.sheetnames:
        if i not in sheets:
            wb2.remove(wb2[i])
    wb2.save('test.xlsx')


    def workMain(dirPath,sheet_name = ''):
        pass







if __name__ == '__main__':

    file = '冠亚广场ODF01-6.xls'
    # sheet_name = '未整理之前'
    # fileContext = getFileContext(file, sheet_name = sheet_name)
    # # print(fileContext)
    # getNewExcelFile(fileContext)

    import pandas as pd
    pd.read_excel(file).to_excel('1.xlsx')