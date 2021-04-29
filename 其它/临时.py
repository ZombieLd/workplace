import openpyxl

def getFileContext(file, sheet_name=''):
    wb = openpyxl.load_workbook(file)
    allSheets = wb.sheetnames
    # print(wb.sheetnames)

    workSheet = wb[sheet_name]

    rows = workSheet.max_row
    data = []
    for i in range(1, rows + 1):
        cell_value = workSheet.cell(row=i, column=7).value
        data.append(cell_value)

    return data




if __name__ == '__main__':
    f = r'C:\Users\asd\Downloads\4月版本提交任务分析(1).xlsx'
    data = getFileContext(f,sheet_name='4月')
    print(data.count('网格党建'))
    print(data.count('心理服务'))
    print(data.count('新版在线采集'))