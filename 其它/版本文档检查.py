# author:liukai
# 用于检查文档：采用清单还是excel

import os,re
import openpyxl
from copy import deepcopy

def getExcelFile(path):
    # 读取清单文档
    all = os.listdir(path)
    listFile = ''
    for filename in all:
        if filename.endswith('xlsx') and '$' not in filename:
            listFile = path + '\\' + filename
    wb = openpyxl.load_workbook(listFile)
    # 获取所有sheets
    sheets = wb.sheetnames
    filelist = []

    for i in range(len(sheets)):
        sheet = wb[sheets[i]]
        rows = sheet.max_row
        col = 2

        for j in range(1, rows):
            row = j + 1
            try:
                filelist.append(sheet.cell(row, col).value.lower())
            except:
                filelist.append(sheet.cell(row, col).value)

    return filelist

def getPathFile(path, releaseNum=''):
    # 获取版本下所有文件
    releaseFiles = []
    releasePath = path + '\\' + releaseNum
    a = os.walk(releasePath)
    for root, dirs, files in a:
        p = files
        # print(p)
        for i in p:
            if not i.startswith('.'):
                releaseFiles.append(i.lower())
    return releaseFiles


def getDiffDoc(path,newVersion,oldVersion):

    exclefirl = getExcelFile(path)
    # print(exclefirl)
    old_release = getPathFile(path, 'release%s'%oldVersion)
    # print(getPathFile(path, 'release20200501'))
    new_release = getPathFile(path, 'release%s'%newVersion)


    # 检查是否提交齐全
    all_doc = deepcopy(old_release)
    # 需要补充到excel的文档
    for i in new_release:
        if (i in old_release) or (i in exclefirl):
            try:
                all_doc.remove(i)
            except:
                print('请手动检查文件：'+i)
        else:
            print('清单需要补充请检查：'+i)
    if all_doc==[]:
        print('清单中文档已都提交！')
    else:
        print("清单中还未提交的文档请检查：")
        for j in all_doc:
            print(j)




if __name__ == '__main__':
    path = r'D:\File\doc\10_相关手册'
    path2 = r'D:\File\doc\30_测试用例'
    getDiffDoc(path2,'20210101','20201201')


